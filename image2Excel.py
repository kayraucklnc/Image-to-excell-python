import sys
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def printRatio(size, percent):
    ratio = int(percent * size)
    text = ["#" for i in range(ratio)] + ["-" for i in range(size - ratio)]
    text = "".join(text)
    print(f"[{text}] %{str(percent)[2:4]}", end="\r")


def processData(imagePath, resulution, outPath):
    image = Image.open(imagePath)
    sizX = image.width / 100 * max(min(resulution, 100), 1)
    sizY = image.height / 100 * max(min(resulution, 100), 1)

    image.thumbnail((sizX, sizY))

    width, height = image.size
    wb = openpyxl.Workbook()

    sheet = wb.active
    counter = 0
    for column in range(1, width):
        printRatio(10, column/width)
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = 2.88
        for row in range(1, height):
            counter = counter + 1
            r, g, b = image.getpixel((column, row))
            hex = '#%02x%02x%02x' % (r, g, b)
            sheet[column_letter + str(row)].fill = PatternFill(
                start_color=hex[1:], end_color=hex[1:], fill_type="solid")
    wb.save(outPath + '.xlsx')


if __name__ == "__main__":
    processData(sys.argv[1], int(sys.argv[2]), sys.argv[3])
